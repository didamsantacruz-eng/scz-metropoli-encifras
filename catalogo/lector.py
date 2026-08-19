# -*- coding: utf-8 -*-
"""
LECTOR GENÉRICO DE LOS TABULADOS DEL INE.
=========================================

Hasta acá había TRES lectores, uno por disposición, cada uno con las columnas
cableadas. No escala: el catálogo declara 57 hojas de validación y cada archivo
del INE arma el encabezado a su manera.

★ LA CLAVE: no adivinar posiciones, LEER LOS RÓTULOS QUE EL PROPIO INE ESCRIBE.
  Todas las hojas traen una fila con `NIVEL · DEPARTAMENTO · PROVINCIA ·
  MUNICIPIO/TIOC · <corte>`, y el corte es SEXO, ÁREA o GRUPO DE EDAD. De ahí
  salen todas las columnas de geografía y dónde empiezan los datos.
  Buscar la palabra "Municipio" en las filas de datos NO sirve: `migracion.xlsx`
  mete una columna `N°` al principio y el rótulo de la cabecera aparece antes,
  así que la detección se corría dos lugares y la hoja quedaba ilegible.

★ EL ENCABEZADO ES UN ÁRBOL de profundidad variable: año → grupo → categoría, o
  año → categoría → Total/Hombres/Mujeres, o sólo año → categoría. Cada columna
  de datos se identifica por su RUTA COMPLETA (la tupla de rótulos que la
  cubren, rellenados hacia la derecha porque las celdas van combinadas). Así una
  misma función sirve para las tres disposiciones sin saber cuál es.

★ LA FILA DEL TOTAL de cada municipio es aquella en la que la columna del corte
  REPITE el nombre del municipio (las otras dicen Hombres/Mujeres, o
  Urbana/Rural). Verificado en 122 de las 132 hojas con municipios; las 10 que
  fallaban eran las de `migracion.xlsx` por el desfase de columnas de arriba.
"""
import pathlib, re, unicodedata
import openpyxl

TAB = pathlib.Path(r"C:\Users\HP\OneDrive\Desktop\Proyectos\Retrato_Censal_2024"
                   r"\Censo2024_Tabulados")
ANIOS = ("2001", "2012", "2024")
IGNORAR = ("Índice", "Glosario", "Ficha técnica")


def norm(s):
    """Minúsculas sin tildes, sin notas al pie y con espacios colapsados."""
    s = str(s if s is not None else "")
    s = re.sub(r"[‐-―−]", " ", s)
    s = unicodedata.normalize("NFD", s).encode("ascii", "ignore").decode()
    s = re.sub(r"\(\d+\)", " ", s)          # notas al pie "(1)"
    s = re.sub(r"[()]", " ", s)
    s = " ".join(s.lower().split())
    # ⚠️ CORRUPCIÓN EN LOS ARCHIVOS DEL INE: `salud.xlsx` escribe
    #    "TIOCaRaqaypampa" —con la LETRA a donde va el guion— y `educacion.xlsx`
    #    escribe "TIOC-Raqaypampa". Sin normalizar el prefijo, los TIOC no cruzan
    #    contra la espina y la hoja aparece con 340 municipios en vez de 343.
    return re.sub(r"^tioc[a ]*", "tioc ", s)


class Hoja:
    """Una hoja del tabulado, ya interpretada."""

    def __init__(self, arch, hoja, filas_crudas):
        self.arch, self.hoja = arch, hoja
        self.error = None
        self.titulo = self._titulo(filas_crudas)
        if not self._geometria(filas_crudas):
            self.error = "no encontré la fila de rótulos NIVEL/DEPARTAMENTO"
            return
        self._encabezado(filas_crudas)
        self._filas(filas_crudas)

    # ---------------------------------------------------------------- montaje
    def _titulo(self, R):
        for f in R[:6]:
            for c in f[:4]:
                t = str(c or "")
                if t.upper().startswith("BOLIVIA"):
                    return " ".join(t.split())
        return ""

    def _geometria(self, R):
        """Ubica las columnas leyendo los rótulos del propio INE."""
        for i, f in enumerate(R[:14]):
            et = {norm(c): j for j, c in enumerate(f) if c}
            if "nivel" not in et or "departamento" not in et:
                continue
            self.rh = i
            self.c_nivel = et["nivel"]
            self.c_dpto = et["departamento"]
            self.c_prov = et.get("provincia", self.c_dpto + 1)
            mun = [j for k, j in et.items() if k.startswith("municipio")]
            self.c_mun = min(mun) if mun else self.c_prov + 1
            # el corte es el rótulo inmediatamente posterior al municipio
            post = sorted(j for j in et.values() if j > self.c_mun)
            self.c_corte = post[0] if post else self.c_mun + 1
            self.dato0 = self.c_corte + 1
            return True
        return False

    def _encabezado(self, R):
        """Ruta completa de cada columna de datos: la tupla de rótulos que la
        cubren, rellenados hacia la derecha (las celdas van combinadas)."""
        ancho = max(len(f) for f in R[:self.rh + 8]) if R else 0
        # ⚠️ El encabezado termina donde EMPIEZAN LOS DATOS. Sin este corte, la
        #    fila de Bolivia (nivel "País") entraba como un nivel más y las rutas
        #    quedaban con el dato pegado al final —('numero de viviendas',
        #    '2024', 'total', '3623711')— con lo que ninguna columna era
        #    reconocible como el total del año.
        NIV = ("pais", "departamento", "provincia", "municipio")
        fin = len(R)
        for i in range(self.rh + 1, len(R)):
            f = R[i]
            if len(f) > self.c_nivel and norm(f[self.c_nivel]).startswith(NIV):
                fin = i
                break
        self.fila0 = fin
        crudas = [f for f in R[self.rh:fin]
                  if any(f[j] not in (None, "") for j in range(self.dato0, min(ancho, len(f))))]
        # ★ El relleno hacia la derecha (celdas combinadas) tiene que REINICIARSE
        #   cuando cambia un nivel de arriba. Rellenando cada fila por separado,
        #   el rótulo de un bloque se filtraba al siguiente y aparecían rutas
        #   imposibles como "… > total > no tiene": la columna del total del año
        #   heredaba el "no tiene" del dispositivo anterior y dejaba de ser
        #   reconocible como denominador.
        niveles = [[""] * ancho for _ in crudas]
        carry = [""] * len(crudas)
        for j in range(ancho):
            for k, f in enumerate(crudas):
                v = f[j] if j < len(f) else None
                if v not in (None, ""):
                    carry[k] = norm(v)
                    for m in range(k + 1, len(crudas)):
                        carry[m] = ""          # el hijo no hereda del bloque previo
                niveles[k][j] = carry[k] if j >= self.dato0 else ""
        self.niveles = niveles
        self.cols = {}
        for j in range(self.dato0, ancho):
            ruta = tuple(n[j] for n in niveles if n[j])
            if ruta:
                self.cols.setdefault(ruta, j)
        self.anios = sorted({int(a) for ruta in self.cols for p in ruta
                             for a in re.findall(r"\b(?:2001|2012|2024)\b", p)})

    def _filas(self, R):
        """Sólo la fila TOTAL de cada municipio (el corte repite su nombre)."""
        self.filas = {}
        for f in R[self.rh + 1:]:
            if len(f) <= self.c_corte:
                continue
            niv = norm(f[self.c_nivel])
            if not niv.startswith("municipio"):
                continue
            mun, corte = norm(f[self.c_mun]), norm(f[self.c_corte])
            if mun and corte == mun:
                self.filas[(norm(f[self.c_dpto]), mun)] = f

    # ---------------------------------------------------------------- consulta
    def columnas(self, anio=None, patrones=(), excluir=("porcent",)):
        """Índices de las columnas cuya ruta case con TODOS los patrones.

        ⚠️ El emparejamiento exige palabra completa: `"tiene"` como subcadena
        también matchea `"NO tiene"`, y con eso el numerador sumaba las dos
        columnas y daba exactamente el denominador (0/343 en agua, electricidad
        y alcantarillado). Sólo se cae a subcadena si no hubo ningún acierto.
        """
        # ★ TRES NIVELES, del más estricto al más laxo, y se devuelve el PRIMERO
        #   que acierte. No es cosmético:
        #   · "tiene" como subcadena también matchea "NO tiene" ⇒ el numerador
        #     sumaba las dos columnas y daba exactamente el denominador.
        #   · "servicio de telefonia fija" es PREFIJO de "servicio de telefonia
        #     fija o telefono celular" ⇒ con prefijo se sumaban las dos, y el
        #     teléfono fijo salía inflado con los celulares.
        def exacto(ruta, p):
            return norm(p) in ruta

        def prefijo(ruta, p):
            p = norm(p)
            return any(x.startswith(p + " ") for x in ruta)

        def cabe_flojo(ruta, p):
            return any(norm(p) in x for x in ruta)

        cand = []
        for ruta, j in self.cols.items():
            txt = " ".join(ruta)
            if anio is not None and str(anio) not in txt:
                continue
            if any(e in txt for e in excluir):
                continue
            cand.append((ruta, j))
        for prueba in (exacto, prefijo, cabe_flojo):
            hit = [j for ruta, j in cand
                   if all(prueba(ruta, p) for p in patrones)]
            if hit:
                return sorted(set(hit))
        return []

    def total(self, anio):
        """La columna 'Total' del año: el denominador que publica el INE.

        Se exige que "total" sea el ÚLTIMO tramo de la ruta. Hojas como `tic/1`
        cuelgan todo de un grupo llamado "Total hogares", así que buscar "total"
        en cualquier posición devolvía las 50 columnas de la hoja.
        """
        cand = []
        for ruta, j in self.cols.items():
            txt = " ".join(ruta)
            if anio is not None and str(anio) not in txt:
                continue
            if "porcent" in txt:
                continue
            # el rótulo del denominador no siempre es "Total" pelado: hay hojas
            # que escriben "Total hogares" o "Total viviendas particulares"
            if ruta[-1] == "total" or ruta[-1].startswith("total "):
                cand.append((len(ruta), j))
        return min(cand)[1] if cand else None

    def categorias(self, anio=None):
        """Rótulos disponibles, para poder mirar qué ofrece la hoja."""
        out = []
        for ruta in self.cols:
            if anio is None or str(anio) in " ".join(ruta):
                out.append(" > ".join(ruta))
        return sorted(set(out))


_cache = {}


def abrir(arch, hoja):
    """Hoja interpretada, con caché (los xlsx del INE son lentos de abrir)."""
    if (arch, hoja) in _cache:
        return _cache[(arch, hoja)]
    ruta = TAB / f"{arch}.xlsx"
    if not ruta.exists():
        _cache[(arch, hoja)] = None
        return None
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    # ⚠️ Los nombres de hoja del INE traen espacios sobrantes: en `ods_pdes` la
    #    hoja se llama literalmente "3.7.2 " con un espacio al final, así que
    #    pedirla como "3.7.2" devolvía "no existe" y el indicador quedaba sin
    #    validar sin que nada avisara. Se busca ignorando espacios.
    real = next((s for s in wb.sheetnames if s.strip() == str(hoja).strip()), None)
    if real is None:
        wb.close(); _cache[(arch, hoja)] = None; return None
    ws = wb[real]
    # ⚠️ En modo read_only openpyxl confía en la dimensión DECLARADA del XML, y
    #    varias hojas del INE la traen mal (`ciudadania`, `emigracion_internacional`,
    #    `movilidad_cotidiana` declaraban una sola columna). `reset_dimensions`
    #    obliga a medir el ancho real al recorrer. Sin esto, 13 hojas se leían
    #    vacías y parecía que tenían otra disposición.
    ws.reset_dimensions()
    R = list(ws.iter_rows(values_only=True))
    wb.close()
    h = Hoja(arch, hoja, R)
    _cache[(arch, hoja)] = h
    return h


def hojas_de(arch):
    ruta = TAB / f"{arch}.xlsx"
    if not ruta.exists():
        return []
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    out = [s for s in wb.sheetnames if s not in IGNORAR]
    wb.close()
    return out


if __name__ == "__main__":
    # mapa de lo que hay: para elegir contra qué validar sin abrir Excel a mano
    import sys
    objetivo = sys.argv[1] if len(sys.argv) > 1 else None
    archivos = [objetivo] if objetivo else sorted(
        p.stem for p in TAB.glob("*.xlsx"))
    tot = ok = 0
    for arch in archivos:
        for hoja in hojas_de(arch):
            h = abrir(arch, hoja)
            tot += 1
            if h is None or h.error:
                print(f"  {arch}/{hoja:<4} ✗ {h.error if h else 'no pude abrir'}")
                continue
            ok += 1
            print(f"  {arch}/{hoja:<4} {len(h.filas):>3} municipios · "
                  f"años {h.anios} · {len(h.cols):>3} columnas · "
                  f"corte col {h.c_corte}")
            if objetivo:
                for c in h.categorias(h.anios[-1] if h.anios else None)[:40]:
                    print(f"        {c}")
    print(f"\n  {ok}/{tot} hojas interpretadas")
